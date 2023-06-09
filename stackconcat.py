from tkinter import filedialog
import pandas as pd
import openpyxl as xl
from openpyxl.styles import PatternFill


def get_input_folder() -> str:
    """Get the file path from the user."""
    file_path = filedialog.askdirectory(initialdir="~/Downloads/")
    return file_path


def add_column(df: any, column_name: str, column_data: str) -> None:
    """Add a column to a dataframe."""
    df[column_name] = column_data


def format_column_as_number(column_number: int) -> None:
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=column_number).number_format = "#,##0.00"


def format_column_as_money(column_number: int) -> None:
    for i in range(2, ws.max_row + 1):
        ws.cell(row=i, column=column_number).number_format = "$#,##0.00"


input_folder = get_input_folder()
job_name = input_folder.split("/")[-1]
df1 = pd.read_excel(input_folder + "/Item Cost By Takeoff.xlsx")
df2 = pd.read_excel(input_folder + "/Takeoff Quantity.xlsx")
output_file = f"{input_folder}/{job_name}_TO.xlsx"

# delete unnecessary columns from df1 and df2
df1 = df1.drop(
    columns=[
        "Accounting Code",
        "Item Name",
        "Item Description",
        "Unit Cost",
        "Cost Type",
        "Extended Cost",
        "Purchase Unit",
    ]
)

del df2["Scale"]

# rename the column in df1
df1.rename(columns={"Takeoff Quantity": "CUFT"}, inplace=True)

# drop all rows with NaN values and reset the index
df1.dropna(inplace=True)
df1.reset_index(drop=True, inplace=True)

# concatenate df1 and df2
df3 = pd.concat([df1, df2], axis=1, join="outer")

# add columns needed for calculations
add_column(df3, "Input Price/Takeoff Unit", "")
add_column(df3, "Total Cost @ Unit Rate", "")
add_column(df3, "Calculated Cost/CUFT", "")
add_column(df3, "Input Price/CUFT", "")
add_column(df3, "Total Cost @ CUFT Rate", "")
add_column(df3, "Calculated Cost/Takeoff Unit", "")
add_column(df3, "Unit", "")

# save dataframe as xlsx file
df3.to_excel(output_file, index=False)

# open the previously output file with openpyxl
wb = xl.load_workbook(output_file)
ws = wb.active

# input formulas for the Total Cost and Calculated columns, color the input cells light orange
for row in range(2, ws.max_row + 1):
    ws.cell(row=row, column=6).fill = xl.styles.PatternFill(
        start_color="FFCC99", end_color="FFCC99", fill_type="solid"
    )
    ws.cell(row=row, column=7).value = f"=F{row}*D{row}"
    ws.cell(row=row, column=8).value = f'=IF(ISBLANK(A{row}),"",G{row}/A{row})'
    ws.cell(row=row, column=9).fill = xl.styles.PatternFill(
        start_color="FFCC99", end_color="FFCC99", fill_type="solid"
    )
    ws.cell(row=row, column=10).value = f"=I{row}*A{row}"
    ws.cell(row=row, column=11).value = f'=IF(ISBLANK(A{row}),"",J{row}/D{row})'
    ws.cell(row=row, column=12).value = f'=IF(ISBLANK(A{row}),"",E{row})'

# input formula to sum columns 7 and 10 in max row + 1
ws.cell(row=ws.max_row + 1, column=7).value = f"=SUM(G2:G{ws.max_row})"
ws.cell(row=ws.max_row, column=10).value = f"=SUM(J2:J{ws.max_row - 1})"

# input "total" one cell to the left of the sum
ws.cell(row=ws.max_row, column=6).value = "Total"
ws.cell(row=ws.max_row, column=9).value = "Total"

# make totals bold
ws.cell(row=ws.max_row, column=6).font = xl.styles.Font(bold=True)
ws.cell(row=ws.max_row, column=9).font = xl.styles.Font(bold=True)

# make the sum cells bold
ws.cell(row=ws.max_row, column=7).font = xl.styles.Font(bold=True)
ws.cell(row=ws.max_row, column=10).font = xl.styles.Font(bold=True)

# add a top border to the total and sum cells
ws.cell(row=ws.max_row, column=6).border = xl.styles.Border(
    top=xl.styles.Side(border_style="thin")
)
ws.cell(row=ws.max_row, column=7).border = xl.styles.Border(
    top=xl.styles.Side(border_style="thin")
)
ws.cell(row=ws.max_row, column=9).border = xl.styles.Border(
    top=xl.styles.Side(border_style="thin")
)
ws.cell(row=ws.max_row, column=10).border = xl.styles.Border(
    top=xl.styles.Side(border_style="thin")
)

number_columns = [1, 4]
money_columns = [6, 7, 8, 9, 10, 11]

for column in number_columns:
    format_column_as_number(column)

for column in money_columns:
    format_column_as_money(column)

# fix all widths to fit the data
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter  # Get the column name
    for cell in col:
        # noinspection PyBroadException
        try:  # Necessary to avoid error on empty cells
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except TypeError:
            pass
    adjusted_width = (max_length + 2) * 1.1
    ws.column_dimensions[column].width = adjusted_width

# save workbook as xlsx file
wb.save(output_file)

print("Done")

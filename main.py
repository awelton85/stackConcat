from tkinter import filedialog
import pandas as pd
import openpyxl as xl
from openpyxl.styles import PatternFill

# open tkinter file dialogs to choose the two Excel files and output path
filename1 = filedialog.askopenfilename(
    initialdir="/home/anthony/Downloads/", title="Select Item Cost By Takeoff"
)
filename2 = filedialog.askopenfilename(
    initialdir="/home/anthony/Downloads/", title="Select Takeoff Quantity"
)
output_file = (
    filedialog.asksaveasfilename(
        initialdir="/home/anthony/Downloads/", title="Select Output File"
    )
    + ".xlsx"
)

df1 = pd.read_excel(filename1)
df2 = pd.read_excel(filename2)

# delete unnecessary columns from df1 and df2
df1 = df1.drop(
    columns=[
        "Accounting Code",
        "Item Name",
        "Item Description",
        "Unit Cost",
        "Cost Type",
        "Extended Cost",
        "Purchase Unit",
    ]
)
del df2["Scale"]

# rename the column in df1
df1.rename(columns={"Takeoff Quantity": "CUFT"}, inplace=True)

# drop all rows with NaN values and reset the index
df1.dropna(inplace=True)
df1.reset_index(drop=True, inplace=True)

# concatenate the two dataframes
df3 = pd.concat([df1, df2], axis=1, join="outer")

# add columns needed for calculations
df3["Input Cost"] = "0"
df3["Total Cost"] = ""
df3["Cost Per CUFT"] = ""

# save dataframe as xlsx file
df3.to_excel(output_file, index=False)

# open the previously output file with openpyxl
wb = xl.load_workbook(output_file)
ws = wb.active

# input formulas for the Total Cost and cost/cuft columns, color the column 6 cells light orange
for row in range(2, ws.max_row + 1):
    ws.cell(row=row, column=7).value = f"=F{row}*D{row}"
    ws.cell(row=row, column=8).value = f'=IF(ISBLANK(A{row}),"",G{row}/A{row})'
    ws.cell(row=row, column=6).fill = xl.styles.PatternFill(
        start_color="FFCC99", end_color="FFCC99", fill_type="solid"
    )

# input formula to sum column 7 in max row + 1
ws.cell(row=ws.max_row + 1, column=7).value = f"=SUM(G2:G{ws.max_row})"

# input "total" one cell to the left of the sum
ws.cell(row=ws.max_row, column=6).value = "Total"

# make total bold
ws.cell(row=ws.max_row, column=6).font = xl.styles.Font(bold=True)

# make the sum cell bold
ws.cell(row=ws.max_row, column=7).font = xl.styles.Font(bold=True)

# add a top border to the total and sum cells
ws.cell(row=ws.max_row, column=6).border = xl.styles.Border(
    top=xl.styles.Side(border_style="thin")
)
ws.cell(row=ws.max_row, column=7).border = xl.styles.Border(
    top=xl.styles.Side(border_style="thin")
)

# fix all widths to fit the data
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter  # Get the column name
    for cell in col:
        try:  # Necessary to avoid error on empty cells
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except Exception as e:
            print(e)
    adjusted_width = (max_length + 2) * 1.2
    ws.column_dimensions[column].width = adjusted_width


# save workbook as xlsx file
wb.save(output_file)
